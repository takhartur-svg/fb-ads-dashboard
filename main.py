from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, FileResponse
from fastapi.templating import Jinja2Templates
import httpx
import io
import csv
import os
from datetime import datetime
import json

app = FastAPI(title="Facebook Ads Dashboard")

# Шукаємо шаблони в різних місцях
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIRS = [
    os.path.join(BASE_DIR, "templates"),
    BASE_DIR,
    "templates",
    "."
]

template_dir = None
for d in TEMPLATE_DIRS:
    if os.path.exists(os.path.join(d, "index.html")):
        template_dir = d
        break

if template_dir:
    templates = Jinja2Templates(directory=template_dir)
else:
    templates = None

# Facebook Graph API base URL
FB_API_URL = "https://graph.facebook.com/v19.0"


class FacebookAdsClient:
    def __init__(self, access_token: str, business_id: str = None, ad_account_id: str = None):
        self.token = access_token
        self.business_id = business_id
        self.ad_account_id = None
        if ad_account_id:
            self.ad_account_id = f"act_{ad_account_id}" if not ad_account_id.startswith("act_") else ad_account_id
    
    async def _request(self, endpoint: str, params: dict = None):
        params = params or {}
        params["access_token"] = self.token
        
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.get(f"{FB_API_URL}/{endpoint}", params=params)
            data = response.json()
            
            if "error" in data:
                raise HTTPException(status_code=400, detail=data["error"].get("message", "Facebook API error"))
            
            return data
    
    def _extract_leads(self, actions):
        """Витягує кількість лідів з actions"""
        if not actions:
            return 0
        for action in actions:
            if action.get("action_type") in ["lead", "onsite_conversion.lead_grouped", "offsite_conversion.fb_pixel_lead"]:
                return int(action.get("value", 0))
        return 0
    
    def _extract_link_clicks(self, actions):
        """Витягує кліки по посиланню"""
        if not actions:
            return 0
        for action in actions:
            if action.get("action_type") == "link_click":
                return int(action.get("value", 0))
        return 0
    
    def _calculate_account_summary(self, insights_data):
        """Розраховує зведені метрики для акаунта"""
        totals = {
            "spend": 0,
            "impressions": 0,
            "reach": 0,
            "clicks": 0,
            "link_clicks": 0,
            "leads": 0
        }
        
        for row in insights_data:
            totals["spend"] += float(row.get("spend", 0))
            totals["impressions"] += int(row.get("impressions", 0))
            totals["reach"] += int(row.get("reach", 0))
            totals["clicks"] += int(row.get("clicks", 0))
            totals["link_clicks"] += self._extract_link_clicks(row.get("actions"))
            totals["leads"] += self._extract_leads(row.get("actions"))
        
        # Розрахунок похідних метрик
        totals["cpm"] = (totals["spend"] / totals["impressions"] * 1000) if totals["impressions"] > 0 else 0
        totals["ctr"] = (totals["clicks"] / totals["impressions"] * 100) if totals["impressions"] > 0 else 0
        totals["cpc"] = (totals["spend"] / totals["clicks"]) if totals["clicks"] > 0 else 0
        totals["cpl"] = (totals["spend"] / totals["leads"]) if totals["leads"] > 0 else 0
        
        return totals

    async def get_business_ad_accounts(self):
        """Отримати всі рекламні акаунти з Business Manager"""
        accounts = []
        endpoint = f"{self.business_id}/owned_ad_accounts"
        params = {"fields": "id,name,currency,account_status,amount_spent", "limit": 100}
        
        while True:
            data = await self._request(endpoint, params)
            accounts.extend(data.get("data", []))
            
            paging = data.get("paging", {})
            if "next" in paging:
                cursors = paging.get("cursors", {})
                if "after" in cursors:
                    params["after"] = cursors["after"]
                else:
                    break
            else:
                break
        
        return accounts
    
    async def get_account_summary(self, account_id: str, account_name: str, currency: str, date_preset: str = "last_30d"):
        """Отримати зведені метрики для одного акаунта"""
        if not account_id.startswith("act_"):
            account_id = f"act_{account_id}"
        
        try:
            data = await self._request(
                f"{account_id}/insights",
                {
                    "fields": "spend,impressions,reach,clicks,actions",
                    "date_preset": date_preset,
                    "limit": 500
                }
            )
            
            insights = data.get("data", [])
            if not insights:
                return None
            
            totals = self._calculate_account_summary(insights)
            totals["account_id"] = account_id
            totals["account_name"] = account_name
            totals["currency"] = currency
            
            return totals
            
        except Exception as e:
            print(f"Error fetching {account_id}: {e}")
            return None
    
    async def get_all_accounts_summary(self, date_preset: str = "last_30d"):
        """Отримати зведені метрики по всіх акаунтах з BM"""
        accounts = await self.get_business_ad_accounts()
        summaries = []
        
        for account in accounts:
            account_id = account["id"]
            account_name = account.get("name", "Unknown")
            currency = account.get("currency", "USD")
            
            summary = await self.get_account_summary(account_id, account_name, currency, date_preset)
            if summary:
                summaries.append(summary)
        
        return {"data": summaries}
    
    async def get_campaign_urls(self):
        """Отримати URL посилань для кампаній"""
        campaign_urls = {}
        
        try:
            # Отримуємо оголошення з креативами
            data = await self._request(
                f"{self.ad_account_id}/ads",
                {
                    "fields": "campaign_id,creative{object_story_spec,asset_feed_spec,link_url}",
                    "limit": 500
                }
            )
            
            for ad in data.get("data", []):
                campaign_id = ad.get("campaign_id")
                creative = ad.get("creative", {})
                
                url = None
                
                # Спробуємо різні місця де може бути URL
                if creative.get("link_url"):
                    url = creative.get("link_url")
                
                object_story_spec = creative.get("object_story_spec", {})
                if object_story_spec:
                    # Link ad
                    link_data = object_story_spec.get("link_data", {})
                    if link_data.get("link"):
                        url = link_data.get("link")
                    
                    # Video ad
                    video_data = object_story_spec.get("video_data", {})
                    if video_data.get("call_to_action", {}).get("value", {}).get("link"):
                        url = video_data["call_to_action"]["value"]["link"]
                
                # Asset feed spec (для динамічних оголошень)
                asset_feed = creative.get("asset_feed_spec", {})
                if asset_feed:
                    link_urls = asset_feed.get("link_urls", [])
                    if link_urls and len(link_urls) > 0:
                        url = link_urls[0].get("website_url")
                
                if url and campaign_id:
                    # Зберігаємо перший знайдений URL для кампанії
                    if campaign_id not in campaign_urls:
                        campaign_urls[campaign_id] = url
                        
        except Exception as e:
            print(f"Error fetching URLs: {e}")
        
        return campaign_urls

    async def get_account_campaigns(self, date_preset: str = "last_30d"):
        """Отримати детальну статистику по кампаніях одного акаунта"""
        data = await self._request(
            f"{self.ad_account_id}/insights",
            {
                "fields": "campaign_name,campaign_id,spend,impressions,reach,clicks,actions,cpm,ctr,cpc",
                "level": "campaign",
                "date_preset": date_preset,
                "limit": 500
            }
        )
        
        # Отримуємо URL для кампаній
        campaign_urls = await self.get_campaign_urls()
        
        campaigns = []
        for row in data.get("data", []):
            campaign_id = row.get("campaign_id")
            campaign = {
                "campaign_name": row.get("campaign_name"),
                "campaign_id": campaign_id,
                "url": campaign_urls.get(campaign_id, ""),
                "spend": float(row.get("spend", 0)),
                "impressions": int(row.get("impressions", 0)),
                "reach": int(row.get("reach", 0)),
                "clicks": int(row.get("clicks", 0)),
                "link_clicks": self._extract_link_clicks(row.get("actions")),
                "leads": self._extract_leads(row.get("actions")),
                "cpm": float(row.get("cpm", 0)),
                "ctr": float(row.get("ctr", 0)),
                "cpc": float(row.get("cpc", 0))
            }
            campaign["cpl"] = (campaign["spend"] / campaign["leads"]) if campaign["leads"] > 0 else 0
            campaigns.append(campaign)
        
        return {"data": campaigns}
    
    async def get_account_daily(self, date_preset: str = "last_14d"):
        """Денна статистика для одного акаунта"""
        data = await self._request(
            f"{self.ad_account_id}/insights",
            {
                "fields": "spend,impressions,clicks,actions,date_start",
                "date_preset": date_preset,
                "time_increment": 1,
                "limit": 500
            }
        )
        
        daily = []
        for row in data.get("data", []):
            daily.append({
                "date_start": row.get("date_start"),
                "spend": float(row.get("spend", 0)),
                "impressions": int(row.get("impressions", 0)),
                "clicks": int(row.get("clicks", 0)),
                "leads": self._extract_leads(row.get("actions"))
            })
        
        return {"data": daily}
    
    async def get_all_accounts_insights(self, date_preset: str = "last_30d"):
        """Отримати статистику по кампаніях всіх акаунтів"""
        accounts = await self.get_business_ad_accounts()
        all_insights = []
        
        for account in accounts:
            account_id = account["id"]
            if not account_id.startswith("act_"):
                account_id = f"act_{account_id}"
            
            try:
                data = await self._request(
                    f"{account_id}/insights",
                    {
                        "fields": "campaign_name,campaign_id,spend,impressions,reach,clicks,actions,cpm,ctr,cpc",
                        "level": "campaign",
                        "date_preset": date_preset,
                        "limit": 500
                    }
                )
                
                for row in data.get("data", []):
                    row["account_name"] = account.get("name", "Unknown")
                    row["account_currency"] = account.get("currency", "USD")
                    row["leads"] = self._extract_leads(row.get("actions"))
                    row["link_clicks"] = self._extract_link_clicks(row.get("actions"))
                    spend = float(row.get("spend", 0))
                    row["cpl"] = (spend / row["leads"]) if row["leads"] > 0 else 0
                    all_insights.append(row)
                    
            except Exception as e:
                print(f"Error fetching {account_id}: {e}")
                continue
        
        return {"data": all_insights}
    
    async def get_all_accounts_daily(self, date_preset: str = "last_14d"):
        """Денна статистика по всіх акаунтах"""
        accounts = await self.get_business_ad_accounts()
        all_insights = []
        
        for account in accounts:
            account_id = account["id"]
            if not account_id.startswith("act_"):
                account_id = f"act_{account_id}"
            
            try:
                data = await self._request(
                    f"{account_id}/insights",
                    {
                        "fields": "spend,impressions,clicks,actions,date_start",
                        "date_preset": date_preset,
                        "time_increment": 1,
                        "limit": 500
                    }
                )
                
                for row in data.get("data", []):
                    row["account_name"] = account.get("name", "Unknown")
                    row["leads"] = self._extract_leads(row.get("actions"))
                    all_insights.append(row)
                    
            except:
                continue
        
        return {"data": all_insights}
    
    async def get_account_info(self):
        if self.ad_account_id:
            return await self._request(
                self.ad_account_id,
                {"fields": "name,currency,account_status,amount_spent,balance"}
            )
        else:
            return await self._request(
                self.business_id,
                {"fields": "name,id"}
            )


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    if templates:
        return templates.TemplateResponse("index.html", {"request": request})
    else:
        for path in ["index.html", "templates/index.html", "app/templates/index.html"]:
            if os.path.exists(path):
                return FileResponse(path, media_type="text/html")
        return HTMLResponse("<h1>index.html not found</h1>", status_code=500)


# === BUSINESS MANAGER ENDPOINTS ===

@app.get("/api/bm/accounts")
async def get_bm_accounts(token: str, business_id: str):
    """Отримати список всіх рекламних акаунтів з BM"""
    client = FacebookAdsClient(token, business_id=business_id)
    accounts = await client.get_business_ad_accounts()
    return {"data": accounts}


@app.get("/api/bm/summary")
async def get_bm_summary(token: str, business_id: str, date_preset: str = "last_30d"):
    """Отримати зведені метрики по кожному акаунту"""
    client = FacebookAdsClient(token, business_id=business_id)
    return await client.get_all_accounts_summary(date_preset=date_preset)


@app.get("/api/bm/insights")
async def get_bm_insights(token: str, business_id: str, date_preset: str = "last_30d"):
    """Отримати статистику по кампаніях всіх акаунтів"""
    client = FacebookAdsClient(token, business_id=business_id)
    return await client.get_all_accounts_insights(date_preset=date_preset)


@app.get("/api/bm/insights/daily")
async def get_bm_daily(token: str, business_id: str, date_preset: str = "last_14d"):
    """Денна статистика"""
    client = FacebookAdsClient(token, business_id=business_id)
    return await client.get_all_accounts_daily(date_preset=date_preset)


@app.get("/api/bm/export/csv")
async def export_bm_csv(token: str, business_id: str, date_preset: str = "last_30d"):
    """Експорт в CSV"""
    client = FacebookAdsClient(token, business_id=business_id)
    data = await client.get_all_accounts_summary(date_preset=date_preset)
    
    output = io.StringIO()
    
    if data.get("data"):
        fieldnames = ["account_name", "spend", "impressions", "leads", "cpl", "reach", "clicks", "link_clicks", "cpm", "ctr", "cpc", "currency"]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in data["data"]:
            writer.writerow({k: row.get(k, "") for k in fieldnames})
    
    output.seek(0)
    filename = f"fb_ads_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/bm/export/excel")
async def export_bm_excel(token: str, business_id: str, date_preset: str = "last_30d"):
    """Експорт в Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    client = FacebookAdsClient(token, business_id=business_id)
    data = await client.get_all_accounts_summary(date_preset=date_preset)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts Summary"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1877F2", end_color="1877F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Акаунт", "Затрати", "Покази", "Ліди", "CPL", "Охвати", "Кліки всі", "Кліки link", "CPM", "CTR", "CPC", "Валюта"]
    fields = ["account_name", "spend", "impressions", "leads", "cpl", "reach", "clicks", "link_clicks", "cpm", "ctr", "cpc", "currency"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    if data.get("data"):
        for row_idx, row_data in enumerate(data["data"], 2):
            for col_idx, field in enumerate(fields, 1):
                value = row_data.get(field, "")
                if isinstance(value, float):
                    value = round(value, 2)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"fb_ads_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# === SINGLE ACCOUNT ENDPOINTS ===

@app.get("/api/account")
async def get_account(token: str, account_id: str):
    client = FacebookAdsClient(token, ad_account_id=account_id)
    return await client.get_account_info()


@app.get("/api/account/summary")
async def get_account_summary(token: str, account_id: str, date_preset: str = "last_30d"):
    """Зведені метрики для одного акаунта"""
    client = FacebookAdsClient(token, ad_account_id=account_id)
    account_info = await client.get_account_info()
    
    summary = await client.get_account_summary(
        account_id, 
        account_info.get("name", "Unknown"),
        account_info.get("currency", "USD"),
        date_preset
    )
    
    return summary or {}


@app.get("/api/account/campaigns")
async def get_account_campaigns(token: str, account_id: str, date_preset: str = "last_30d"):
    """Статистика по кампаніях акаунта"""
    client = FacebookAdsClient(token, ad_account_id=account_id)
    return await client.get_account_campaigns(date_preset=date_preset)


@app.get("/api/account/daily")
async def get_account_daily(token: str, account_id: str, date_preset: str = "last_14d"):
    """Денна статистика акаунта"""
    client = FacebookAdsClient(token, ad_account_id=account_id)
    return await client.get_account_daily(date_preset=date_preset)


@app.get("/api/account/export/csv")
async def export_account_csv(token: str, account_id: str, date_preset: str = "last_30d"):
    """Експорт кампаній в CSV"""
    client = FacebookAdsClient(token, ad_account_id=account_id)
    data = await client.get_account_campaigns(date_preset=date_preset)
    
    output = io.StringIO()
    
    if data.get("data"):
        fieldnames = ["campaign_name", "url", "spend", "impressions", "leads", "cpl", "reach", "clicks", "link_clicks", "cpm", "ctr", "cpc"]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in data["data"]:
            writer.writerow({k: row.get(k, "") for k in fieldnames})
    
    output.seek(0)
    filename = f"fb_campaigns_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/account/export/excel")
async def export_account_excel(token: str, account_id: str, date_preset: str = "last_30d"):
    """Експорт кампаній в Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    client = FacebookAdsClient(token, ad_account_id=account_id)
    data = await client.get_account_campaigns(date_preset=date_preset)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campaigns"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1877F2", end_color="1877F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Кампанія", "URL", "Затрати", "Покази", "Ліди", "CPL", "Охвати", "Кліки всі", "Кліки link", "CPM", "CTR", "CPC"]
    fields = ["campaign_name", "url", "spend", "impressions", "leads", "cpl", "reach", "clicks", "link_clicks", "cpm", "ctr", "cpc"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    if data.get("data"):
        for row_idx, row_data in enumerate(data["data"], 2):
            for col_idx, field in enumerate(fields, 1):
                value = row_data.get(field, "")
                if isinstance(value, float):
                    value = round(value, 2)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # URL column wider
        ws.column_dimensions['B'].width = 40
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"fb_campaigns_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
